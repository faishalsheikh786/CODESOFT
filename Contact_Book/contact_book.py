import tkinter as tk
from tkinter import ttk
import os
import openpyxl

def load_data():
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Arial', 12, 'bold'))
    style.configure("Treeview", font=('Arial', 10   ))
    path = "people.xlsx"
    if not os.path.exists(path):
        # Create a new workbook and add a sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Add headers to the sheet
        headers = ["Name", "Contact", "Email", "Address"]
        sheet.append(headers)
        # Save the workbook
        workbook.save(path)

    tree_view.delete(*tree_view.get_children())
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    
    for col_name in list_values[0]:
        tree_view.heading(col_name, text=col_name, anchor="center")
        tree_view.column(col_name, anchor="center")

    for value_tuple in list_values[1:]:
        tree_view.insert('', tk.END, values=value_tuple)


def enable_buttons():
    button1.config(state="normal")
    button2.config(state="normal")
    button3.config(state="normal")
    button4.config(state="normal")


def disable_buttons():
    button1.config(state="disabled")
    button2.config(state="disabled")
    button3.config(state="disabled")
    button4.config(state="disabled")


def addData():
    disable_buttons()

    def focusin_nameEntry(event):
        name.delete(0, tk.END)

    def focusout_nameEntry(event):
        if name.get()=="":
            name.insert(0, "Name")

    def focusin_contactEntry(event):
        contact.delete(0, tk.END)

    def focusout_contactEntry(event):
        if contact.get()=="":
            contact.insert(0, "Contact")

    def focusin_emailEntry(event):
        email.delete(0, tk.END)

    def focusout_emailEntry(event):
        if email.get()=="":
            email.insert(0, "Email")

    def focusin_addressEntry(event):
        address.delete(0, tk.END)

    def focusout_addressEntry(event):
        if address.get()=="":
            address.insert(0, "Address")

    def save_data():
        path = "people.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [name.get(), contact.get(), email.get(), address.get()] 
        sheet.append(row_values)
        workbook.save(path)
        tree_view.insert('', tk.END, values=row_values)
        enable_buttons()
        addData_window.destroy()

    addData_window = tk.Tk()
    addData_window.title("Add Contact")
    addData_window.iconbitmap("notebook-of-contacts.ico")

    name = tk.Entry(addData_window, width=50, font=('Arial', 18), highlightcolor='green')
    name.insert(0, "Name")
    name.bind("<FocusIn>", focusin_nameEntry)
    name.bind("<FocusOut>", focusout_nameEntry)
    name.grid(row=0, column=0, padx=10, pady=10)

    contact = tk.Entry(addData_window, width=50, font=('Arial', 18))
    contact.insert(0, "Contact")
    contact.bind("<FocusIn>", focusin_contactEntry)
    contact.bind("<FocusOut>", focusout_contactEntry)
    contact.grid(row=1, column=0, pady=10)

    email = tk.Entry(addData_window, width=50, font=('Arial', 18))
    email.insert(0, "Email")
    email.bind("<FocusIn>", focusin_emailEntry)
    email.bind("<FocusOut>", focusout_emailEntry)
    email.grid(row=2, column=0, pady=10)

    address = tk.Entry(addData_window, width=50, font=('Arial', 18))
    address.insert(0, "Address")
    address.bind("<FocusIn>", focusin_addressEntry)
    address.bind("<FocusOut>", focusout_addressEntry)
    address.grid(row=3, column=0, pady=10)

    save_button = tk.Button(addData_window, text="Save", font=('Arial', 18), command=save_data)
    save_button.grid(row=4, column=0, pady=10)

    def on_closing():
        addData_window.destroy()
        enable_buttons()
    addData_window.protocol("WM_DELETE_WINDOW", on_closing)

    addData_window.mainloop()



def updateData():

    disable_buttons()
    
    def save_data():
        path = "people.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == selected_item[0] or row[1] == selected_item[1] or row[2] == selected_item[2] or row[3] == selected_item[3]:
                sheet.cell(row=row_index, column=1, value=name.get())  # Update Name
                sheet.cell(row=row_index, column=2, value=contact.get())  # Update Contact
                sheet.cell(row=row_index, column=3, value=email.get())    # Update Email
                sheet.cell(row=row_index, column=4, value=address.get())  # Update Address
                break
        workbook.save(path)
        load_data()
        enable_buttons()
        updateData_window.destroy()
    
    if not tree_view.selection():

        def destroyOk():
            enable_buttons()
            updateData_window.destroy()

        updateData_window = tk.Tk()
        updateData_window.title("Alert")
        updateData_window.iconbitmap("notebook-of-contacts.ico")

        label = tk.Label(updateData_window, text="First Select the Contact", font=("Arial", 18))
        label.pack(padx=10, pady=10)

        button = tk.Button(updateData_window, text='Ok', font=("Arial", 18), command=destroyOk)
        button.pack(padx=10, pady=10)

        def on_closing():
            updateData_window.destroy()
            enable_buttons()
        updateData_window.protocol("WM_DELETE_WINDOW", on_closing)
        updateData_window.mainloop()

    else:
        updateData_window = tk.Tk()
        updateData_window.title("Update Contact")
        updateData_window.iconbitmap("notebook-of-contacts.ico")

        name = tk.Entry(updateData_window, width=50, font=('Arial', 18))
        name.insert(0, selected_item[0])
        name.grid(row=0, column=0, padx=10, pady=10)

        contact = tk.Entry(updateData_window, width=50, font=('Arial', 18))
        contact.insert(0, selected_item[1])
        contact.grid(row=1, column=0, pady=10)

        email = tk.Entry(updateData_window, width=50, font=('Arial', 18))
        email.insert(0, selected_item[2])
        email.grid(row=2, column=0, pady=10)

        address = tk.Entry(updateData_window, width=50, font=('Arial', 18))
        address.insert(0, selected_item[3])
        address.grid(row=3, column=0, pady=10)

        save_button = tk.Button(updateData_window, text="Save", font=('Arial', 18), command=save_data)
        save_button.grid(row=4, column=0, pady=10)

        def on_closing():
            updateData_window.destroy()
            enable_buttons()
        updateData_window.protocol("WM_DELETE_WINDOW", on_closing)

        updateData_window.mainloop()



def deleteData():

    disable_buttons()
    
    def delete_data():
        selected_item = tree_view.selection()
        if selected_item:
            # Get the index of the selected row in the TreeView
            row_index = tree_view.index(selected_item)

            # Open the workbook
            path = "people.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active

            # Delete the row from the worksheet
            sheet.delete_rows(row_index + 2)  # Add 1 to row_index to match Excel row index

            # Save the workbook
            workbook.save(path)

            # Reload data into the TreeView
            load_data()
            enable_buttons()
            deleteData_window.destroy()

    if not tree_view.selection():

        def destroyOk():
            enable_buttons()
            deleteData_window.destroy()

        deleteData_window = tk.Tk()
        deleteData_window.title("Alert")
        deleteData_window.iconbitmap("notebook-of-contacts.ico")

        label = tk.Label(deleteData_window, text="First Select the Contact", font=("Arial", 18))
        label.pack(padx=10, pady=10)

        button = tk.Button(deleteData_window, text='Ok', font=("Arial", 18), command=destroyOk)
        button.pack(padx=10, pady=10)

        def on_closing():
            deleteData_window.destroy()
            enable_buttons()
        deleteData_window.protocol("WM_DELETE_WINDOW", on_closing)

        deleteData_window.mainloop()

    else:
        deleteData_window = tk.Tk()
        deleteData_window.minsize(height=200, width=300)
        deleteData_window.title("Delete Contact")
        deleteData_window.iconbitmap("notebook-of-contacts.ico")

        name_label = tk.Label(deleteData_window, text=f"Name : {selected_item[0]}", font=("Arial", 18))
        contact_label = tk.Label(deleteData_window, text=f"Contact : {selected_item[1]}", font=("Arial", 18))
        email_label = tk.Label(deleteData_window, text=f"Email : {selected_item[2]}", font=("Arial", 18))
        address_label = tk.Label(deleteData_window, text=f"Address : {selected_item[3]}", font=("Arial", 18))

        name_label.pack(padx=10, pady=5)
        contact_label.pack(padx=10, pady=5)
        email_label.pack(padx=10, pady=5)
        address_label.pack(padx=10, pady=5)

        delete_button = tk.Button(deleteData_window, text="Delete", font=("Arial", 18), command=delete_data)
        delete_button.pack(padx=10, pady=20)

        def on_closing():
            deleteData_window.destroy()
            enable_buttons()
        deleteData_window.protocol("WM_DELETE_WINDOW", on_closing)

        deleteData_window.mainloop()


def searchData():
    search_text = search_entry.get().lower()
    path = "people.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Clear existing data in the TreeView
    tree_view.delete(*tree_view.get_children())

    # Iterate through rows and filter based on search criteria
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if search_text in (str(cell).lower() for cell in row):
            tree_view.insert("", tk.END, values=row)

    search_entry.delete(0, tk.END)
    root.focus_set()


def on_treeview_focus_in():
    selection = tree_view.selection()
    if selection:
        global selected_item
        selected_item = tree_view.item(selection, "values")
        # print("Selected value:", selected_item)
    else:
        pass
    root.after(100, on_treeview_focus_in)



root = tk.Tk()
root.title("Contact Book")
root.iconbitmap("notebook-of-contacts.ico")

frame1 = tk.Frame(root)
frame1.grid(row=0, column=0, padx=20, pady=20)

button1 = tk.Button(frame1, text='Add', font=('Arial',15),height=2, width=8, command=addData)
button2 = tk.Button(frame1, text='View', font=('Arial',15),height=2, width=8, command=load_data)
button3 = tk.Button(frame1, text='Update', font=('Arial',15),height=2, width=8, command=updateData)
button4 = tk.Button(frame1, text='Delete', font=('Arial',15),height=2, width=8, command=deleteData)

button1.pack(pady=10)
button2.pack(pady=10)
button3.pack(pady=10)
button4.pack(pady=10)

frame2 = tk.Frame(root)
frame2.grid(row=0, column=1)

frame2_1 = tk.Frame(frame2)
frame2_1.grid(row=0, column=0)

search_entry = tk.Entry(frame2_1, width=50)
search_entry.grid(row=0, column=0, padx=10)

search_button = tk.Button(frame2_1, text='Search', font=('Arial', 10, 'bold'), command=searchData)
search_button.grid(row=0, column=1)

frame2_2 = tk.Frame(frame2)
frame2_2.grid(row=1, column=0, pady=10)
treeScrollY = ttk.Scrollbar(frame2_2)
treeScrollY.pack(side="right", fill="y")

cols = ("Name", "Contact", "Email", "Address")
tree_view = ttk.Treeview(frame2_2, show='headings', columns=cols, height=13)
tree_view.column("Name", width=250)
tree_view.column("Contact", width=150)
tree_view.column("Email", width=350)
tree_view.column("Address", width=500)
tree_view.bind("<FocusIn>", lambda event:on_treeview_focus_in())
tree_view.bind("<FocusOut>", lambda event:on_treeview_focus_in())
tree_view.pack()
treeScrollY.config(command=tree_view.yview)
load_data()

root.mainloop()



